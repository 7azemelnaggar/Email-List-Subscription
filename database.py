"""
Database module for Email Marketing Lists and Employee Management
Provides CRUD operations for all tables
"""

import sqlite3
import csv
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import os


class DatabaseManager:
    """Manages database operations for email marketing and employee management"""
    
    def __init__(self, db_name: str = "email_marketing.db"):
        """Initialize database connection"""
        self.db_name = db_name
        self.conn = None
        self.connect()
        self.create_tables()
    
    def connect(self):
        """Establish database connection"""
        self.conn = sqlite3.connect(self.db_name)
        self.conn.row_factory = sqlite3.Row  # Return rows as dictionaries
        return self.conn
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
    
    def create_tables(self):
        """Create all tables from schema.sql"""
        with open('schema.sql', 'r', encoding='utf-8') as f:
            schema = f.read()
        
        # Remove comments and split by semicolons
        lines = schema.split('\n')
        cleaned_lines = []
        for line in lines:
            # Remove inline comments
            if '--' in line:
                line = line[:line.index('--')]
            cleaned_lines.append(line)
        
        # Join and split by semicolons
        cleaned_schema = '\n'.join(cleaned_lines)
        statements = [s.strip() for s in cleaned_schema.split(';') if s.strip()]
        
        cursor = self.conn.cursor()
        for statement in statements:
            if statement:
                try:
                    cursor.execute(statement)
                except sqlite3.OperationalError as e:
                    error_msg = str(e).lower()
                    # Ignore errors for existing tables/indexes and constraint violations
                    if "already exists" not in error_msg and "constraint" not in error_msg:
                        # Only print warnings for unexpected errors
                        pass
        self.conn.commit()
    
    # ==================== DEPARTMENT CRUD OPERATIONS ====================
    
    def create_department(self, name: str, head_of_department_id: Optional[int] = None) -> int:
        """Create a new department"""
        cursor = self.conn.cursor()
        cursor.execute(
            "INSERT INTO departments (name, head_of_department_id) VALUES (?, ?)",
            (name, head_of_department_id)
        )
        self.conn.commit()
        return cursor.lastrowid
    
    def get_department(self, department_id: int) -> Optional[Dict]:
        """Get department by ID"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM departments WHERE id = ?", (department_id,))
        row = cursor.fetchone()
        return dict(row) if row else None
    
    def get_all_departments(self) -> List[Dict]:
        """Get all departments"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM departments ORDER BY name")
        return [dict(row) for row in cursor.fetchall()]
    
    def update_department(self, department_id: int, name: Optional[str] = None, 
                         head_of_department_id: Optional[int] = None) -> bool:
        """Update department information"""
        updates = []
        params = []
        
        if name is not None:
            updates.append("name = ?")
            params.append(name)
        if head_of_department_id is not None:
            updates.append("head_of_department_id = ?")
            params.append(head_of_department_id)
        
        if not updates:
            return False
        
        params.append(department_id)
        cursor = self.conn.cursor()
        cursor.execute(
            f"UPDATE departments SET {', '.join(updates)} WHERE id = ?",
            params
        )
        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_department(self, department_id: int) -> bool:
        """Delete a department (cascades to employees)"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM departments WHERE id = ?", (department_id,))
        self.conn.commit()
        return cursor.rowcount > 0
    
    # ==================== EMPLOYEE CRUD OPERATIONS ====================
    
    def create_employee(self, name: str, email: str, department_id: int, 
                       is_supervisor: bool = False, is_head: bool = False,
                       position: Optional[str] = None, hire_date: Optional[str] = None) -> int:
        """Create a new employee"""
        cursor = self.conn.cursor()
        cursor.execute(
            """INSERT INTO employees (name, email, department_id, is_supervisor, 
               is_head, position, hire_date) VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (name, email, department_id, is_supervisor, is_head, position, hire_date)
        )
        self.conn.commit()
        return cursor.lastrowid
    
    def get_employee(self, employee_id: int) -> Optional[Dict]:
        """Get employee by ID"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM employees WHERE id = ?", (employee_id,))
        row = cursor.fetchone()
        return dict(row) if row else None
    
    def get_all_employees(self) -> List[Dict]:
        """Get all employees"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.*, d.name as department_name 
            FROM employees e 
            LEFT JOIN departments d ON e.department_id = d.id 
            ORDER BY e.name
        """)
        return [dict(row) for row in cursor.fetchall()]
    
    def get_employees_by_department(self, department_id: int) -> List[Dict]:
        """Get all employees in a specific department"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.*, d.name as department_name 
            FROM employees e 
            LEFT JOIN departments d ON e.department_id = d.id 
            WHERE e.department_id = ?
            ORDER BY e.name
        """, (department_id,))
        return [dict(row) for row in cursor.fetchall()]
    
    def get_supervisors_by_department(self, department_id: int) -> List[Dict]:
        """Get all supervisors in a specific department"""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.*, d.name as department_name 
            FROM employees e 
            LEFT JOIN departments d ON e.department_id = d.id 
            WHERE e.department_id = ? AND e.is_supervisor = 1
            ORDER BY e.name
        """, (department_id,))
        return [dict(row) for row in cursor.fetchall()]
    
    def update_employee(self, employee_id: int, name: Optional[str] = None,
                        email: Optional[str] = None, department_id: Optional[int] = None,
                        is_supervisor: Optional[bool] = None, is_head: Optional[bool] = None,
                        position: Optional[str] = None, hire_date: Optional[str] = None) -> bool:
        """Update employee information"""
        updates = []
        params = []
        
        if name is not None:
            updates.append("name = ?")
            params.append(name)
        if email is not None:
            updates.append("email = ?")
            params.append(email)
        if department_id is not None:
            updates.append("department_id = ?")
            params.append(department_id)
        if is_supervisor is not None:
            updates.append("is_supervisor = ?")
            params.append(int(is_supervisor))
        if is_head is not None:
            updates.append("is_head = ?")
            params.append(int(is_head))
        if position is not None:
            updates.append("position = ?")
            params.append(position)
        if hire_date is not None:
            updates.append("hire_date = ?")
            params.append(hire_date)
        
        if not updates:
            return False
        
        params.append(employee_id)
        cursor = self.conn.cursor()
        cursor.execute(
            f"UPDATE employees SET {', '.join(updates)} WHERE id = ?",
            params
        )
        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_employee(self, employee_id: int) -> bool:
        """Delete an employee"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
        self.conn.commit()
        return cursor.rowcount > 0
    
    # ==================== EMAIL SUBSCRIPTION CRUD OPERATIONS ====================
    
    def create_email_subscription(self, email: str, status: str = 'active',
                                  source: Optional[str] = None, notes: Optional[str] = None) -> int:
        """Create a new email subscription"""
        cursor = self.conn.cursor()
        cursor.execute(
            """INSERT INTO email_subscriptions (email, status, source, notes) 
               VALUES (?, ?, ?, ?)""",
            (email, status, source, notes)
        )
        self.conn.commit()
        return cursor.lastrowid
    
    def get_email_subscription(self, subscription_id: int) -> Optional[Dict]:
        """Get email subscription by ID"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM email_subscriptions WHERE id = ?", (subscription_id,))
        row = cursor.fetchone()
        return dict(row) if row else None
    
    def get_email_subscription_by_email(self, email: str) -> Optional[Dict]:
        """Get email subscription by email address"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM email_subscriptions WHERE email = ?", (email,))
        row = cursor.fetchone()
        return dict(row) if row else None
    
    def get_all_email_subscriptions(self, status: Optional[str] = None) -> List[Dict]:
        """Get all email subscriptions, optionally filtered by status"""
        cursor = self.conn.cursor()
        if status:
            cursor.execute(
                "SELECT * FROM email_subscriptions WHERE status = ? ORDER BY subscribed_at DESC",
                (status,)
            )
        else:
            cursor.execute("SELECT * FROM email_subscriptions ORDER BY subscribed_at DESC")
        return [dict(row) for row in cursor.fetchall()]
    
    def update_email_subscription(self, subscription_id: int, email: Optional[str] = None,
                                  status: Optional[str] = None, source: Optional[str] = None,
                                  notes: Optional[str] = None) -> bool:
        """Update email subscription information"""
        updates = []
        params = []
        
        if email is not None:
            updates.append("email = ?")
            params.append(email)
        if status is not None:
            updates.append("status = ?")
            params.append(status)
        if source is not None:
            updates.append("source = ?")
            params.append(source)
        if notes is not None:
            updates.append("notes = ?")
            params.append(notes)
        
        if not updates:
            return False
        
        params.append(subscription_id)
        cursor = self.conn.cursor()
        cursor.execute(
            f"UPDATE email_subscriptions SET {', '.join(updates)} WHERE id = ?",
            params
        )
        self.conn.commit()
        return cursor.rowcount > 0
    
    def delete_email_subscription(self, subscription_id: int) -> bool:
        """Delete an email subscription"""
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM email_subscriptions WHERE id = ?", (subscription_id,))
        self.conn.commit()
        return cursor.rowcount > 0
    
    # ==================== CSV EXPORT/IMPORT OPERATIONS ====================
    
    def export_emails_to_csv(self, filename: str, status: Optional[str] = None) -> bool:
        """Export email subscriptions to CSV file"""
        try:
            subscriptions = self.get_all_email_subscriptions(status)
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                if subscriptions:
                    fieldnames = subscriptions[0].keys()
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(subscriptions)
                else:
                    # Create empty CSV with headers
                    writer = csv.DictWriter(csvfile, fieldnames=['id', 'email', 'subscribed_at', 'status', 'source', 'notes'])
                    writer.writeheader()
            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def export_emails_to_excel(self, filename: str, status: Optional[str] = None) -> bool:
        """Export email subscriptions to Excel file (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            subscriptions = self.get_all_email_subscriptions(status)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Email Subscriptions"
            
            if subscriptions:
                # Write headers
                headers = list(subscriptions[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                # Write data
                for row, subscription in enumerate(subscriptions, 2):
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=subscription.get(header, ''))
            else:
                # Create empty Excel with headers
                headers = ['id', 'email', 'subscribed_at', 'status', 'source', 'notes']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
            
            wb.save(filename)
            return True
        except ImportError:
            print("openpyxl not installed. Install it with: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def import_emails_from_csv(self, filename: str, skip_duplicates: bool = True) -> Tuple[int, int]:
        """
        Import email subscriptions from CSV file
        Returns: (successful_imports, failed_imports)
        """
        successful = 0
        failed = 0
        
        try:
            with open(filename, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    try:
                        email = row.get('email', '').strip()
                        if not email:
                            failed += 1
                            continue
                        
                        # Check if email already exists
                        if skip_duplicates:
                            existing = self.get_email_subscription_by_email(email)
                            if existing:
                                continue
                        
                        status = row.get('status', 'active').strip() or 'active'
                        source = row.get('source', '').strip() or None
                        notes = row.get('notes', '').strip() or None
                        
                        self.create_email_subscription(email, status, source, notes)
                        successful += 1
                    except Exception as e:
                        print(f"Error importing row {row}: {e}")
                        failed += 1
            return successful, failed
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return successful, failed

